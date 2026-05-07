"""
마스터 데이터 import 스크립트
- 5개 엑셀 파일 → Supabase 테이블 5개 (products/vendors/materials/bom/drawings)
- 거래 ledger도 동시 import (sales/purchase)
- products는 자동집계 28개 컬럼 제외 (정적 정보만)
- 휴면 3년+ 제품은 archived_at 자동 채움

[사용 방법] Streamlit 앱의 "마스터 관리" 화면에서 "import 실행" 버튼 클릭.
이 스크립트는 streamlit_app.py의 함수로 호출됨.
"""
import os
import re
from datetime import datetime, timedelta
import openpyxl
from supabase import Client


# ────────────────────────────────────────────────
# 정적 컬럼만 추출하는 매핑
# ────────────────────────────────────────────────

PRODUCT_COL_MAP = {
    'ID': 'product_id',
    '품번': 'pn',
    'alias': 'alias_list',
    '도면번호': 'drawing_no',
    '소분류': 'sub_class',
    '제품군': 'product_group',
    '재질': 'material',
    '조달유형': 'procurement_type',
    '조달유형_시작일': 'procurement_start_date',
    '이전_조달유형': 'procurement_prev_type',
    '원자재명': 'raw_material_name',
    '소재규격': 'raw_material_spec',
    '거래처': 'customer',
    'BOM_자재명': 'bom_material_name',
    '자재_단위': 'material_unit',
    '자재_KG단가': 'material_kg_price',
    '자재_개당단가': 'material_unit_price',
    '자재_매입건수': 'material_purchase_count',
    '자재_최근매입일': 'material_last_purchase_date',
    '자재_주공급사': 'material_main_supplier',
    '자재_데이터품질': 'material_data_quality',
    '외주가공비_per_pc': 'outsourcing_per_pc',
    '열처리비_per_pc': 'heat_treat_per_pc',
    '표면처리비_per_pc': 'surface_per_pc',
    '추정원가_per_pc': 'estimated_cost_per_pc',
    '원가데이터품질': 'cost_data_quality',
    '활성': 'active',
    '주의': 'caution',
    '추정근거': 'inference_basis',
}

VENDOR_COL_MAP = {
    '거래처명': 'name',
    '정규화명': 'normalized_name',
    '사업자번호': 'business_no',
    '거래처 구분': 'trade_type',
    '카테고리': 'category',
    '대표자명': 'ceo_name',
    '전화번호': 'phone',
    'Fax번호': 'fax',
    '주소': 'address',
    '이메일': 'email',
    '담당자명': 'contact_person',
    '담당자연락처': 'contact_phone',
    '업태': 'business_type',
    '종목': 'business_item',
    '결제조건': 'payment_terms',
    '메모사항': 'memo',
    '확인상태': 'verification_status',
}

MATERIAL_COL_MAP = {
    '자재ID': 'material_id',
    '원자재명': 'raw_name',
    '재질': 'material_type',
    '규격': 'spec',
    '조달유형': 'procurement_type',
    '재고량': 'stock_qty',
    '사용여부': 'in_use',
    '주 공급사(매입이력 기준)': 'main_supplier',
}

BOM_COL_MAP = {
    'product_id': 'product_id',
    '대표품번': 'pn',
    '자재ID': 'material_id',
    '원자재명': 'raw_material_name',
    '소요량/PC': 'qty_per_pc',
    '출처': 'source',
    '매칭근거': 'match_basis',
    '적용_시작일': 'apply_start_date',
    '적용_종료일': 'apply_end_date',
    '확인상태': 'verification_status',
}


def to_text(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None


def to_num(v):
    if v is None or v == '': return None
    try: return float(str(v).replace(',', ''))
    except: return None


def to_int(v):
    n = to_num(v)
    return int(n) if n is not None else None


def to_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date().isoformat()
    s = str(v).strip()
    if not s: return None
    # YYYY-MM 또는 YYYY-MM-DD 형식 시도
    for fmt in ('%Y-%m-%d', '%Y-%m', '%Y/%m/%d'):
        try: return datetime.strptime(s, fmt).date().isoformat()
        except: pass
    return None


def read_xlsx(path: str, sheet: str = None) -> tuple[list[str], list[list]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append([ws.cell(r, ci).value for ci in range(1, len(headers) + 1)])
    return headers, rows


def map_row(headers: list, row: list, col_map: dict) -> dict:
    out = {}
    for src, dst in col_map.items():
        if src in headers:
            v = row[headers.index(src)]
            out[dst] = v
    return out


def import_products(client: Client, src_dir: str) -> dict:
    """product_master_v11.xlsx → products 테이블 (정적 컬럼만)"""
    path = os.path.join(src_dir, 'product_master_v11.xlsx')
    headers, rows = read_xlsx(path, 'product_master')

    records = []
    for row in rows:
        d = map_row(headers, row, PRODUCT_COL_MAP)
        if not d.get('pn'): continue
        # 타입 변환
        for k in ('material_kg_price','material_unit_price','outsourcing_per_pc','heat_treat_per_pc',
                  'surface_per_pc','estimated_cost_per_pc'):
            d[k] = to_num(d.get(k))
        d['material_purchase_count'] = to_int(d.get('material_purchase_count'))
        d['procurement_start_date'] = to_date(d.get('procurement_start_date'))
        for k in ('product_id','pn','alias_list','drawing_no','sub_class','product_group','material',
                  'procurement_type','procurement_prev_type','raw_material_name','raw_material_spec',
                  'customer','bom_material_name','material_unit','material_last_purchase_date',
                  'material_main_supplier','material_data_quality','cost_data_quality',
                  'active','caution','inference_basis'):
            d[k] = to_text(d.get(k))
        if not d.get('product_id'):
            d['product_id'] = d['pn']  # fallback
        records.append(d)

    # upsert
    BATCH = 500
    inserted = 0
    for i in range(0, len(records), BATCH):
        chunk = records[i:i+BATCH]
        client.table('products').upsert(chunk).execute()
        inserted += len(chunk)
    return {'table': 'products', 'count': inserted}


def import_vendors(client: Client, src_dir: str) -> dict:
    path = os.path.join(src_dir, '거래처관리_v2.xlsx')
    headers, rows = read_xlsx(path, '거래처')
    records = []
    for row in rows:
        d = map_row(headers, row, VENDOR_COL_MAP)
        if not d.get('name'): continue
        for k in ('name','normalized_name','business_no','trade_type','category','ceo_name',
                  'phone','fax','address','email','contact_person','contact_phone',
                  'business_type','business_item','payment_terms','memo','verification_status'):
            d[k] = to_text(d.get(k))
        records.append(d)

    BATCH = 500
    for i in range(0, len(records), BATCH):
        client.table('vendors').upsert(records[i:i+BATCH], on_conflict='business_no').execute()
    return {'table': 'vendors', 'count': len(records)}


def import_materials(client: Client, src_dir: str) -> dict:
    path = os.path.join(src_dir, 'material_master_v3.xlsx')
    headers, rows = read_xlsx(path, 'material_master')
    records = []
    for row in rows:
        d = map_row(headers, row, MATERIAL_COL_MAP)
        if not d.get('material_id'): continue
        d['stock_qty'] = to_num(d.get('stock_qty')) or 0
        for k in ('material_id','raw_name','material_type','spec','procurement_type',
                  'in_use','main_supplier'):
            d[k] = to_text(d.get(k))
        records.append(d)
    BATCH = 500
    for i in range(0, len(records), BATCH):
        client.table('materials').upsert(records[i:i+BATCH]).execute()
    return {'table': 'materials', 'count': len(records)}


def import_bom(client: Client, src_dir: str) -> dict:
    path = os.path.join(src_dir, 'BOM_v8.xlsx')
    headers, rows = read_xlsx(path, 'BOM')
    records = []
    for row in rows:
        d = map_row(headers, row, BOM_COL_MAP)
        # PN을 product_id로 변환은 별도 로직 필요. 우선 그대로 기록
        if not d.get('product_id'): continue
        d['qty_per_pc'] = to_num(d.get('qty_per_pc')) or 1
        d['apply_start_date'] = to_date(d.get('apply_start_date'))
        d['apply_end_date'] = to_date(d.get('apply_end_date'))
        for k in ('product_id','material_id','raw_material_name','source','match_basis','verification_status'):
            d[k] = to_text(d.get(k))
        # pn은 별도 컬럼이 아님 (조회용) → 제거
        d.pop('pn', None)
        records.append(d)
    BATCH = 500
    inserted = 0
    for i in range(0, len(records), BATCH):
        client.table('bom').insert(records[i:i+BATCH]).execute()
        inserted += len(records[i:i+BATCH])
    return {'table': 'bom', 'count': inserted}


def import_drawings(client: Client, src_dir: str) -> dict:
    path = os.path.join(src_dir, '도면관리.xlsx')
    headers, rows = read_xlsx(path, '제품목록')
    DRAW_MAP = {
        '고객사': 'customer', '품번': 'pn', '리비전': 'revision',
        '리비전상태': 'revision_status', '파일명': 'filename',
        '확장자': 'file_ext', '크기(KB)': 'file_size_kb',
        '수정일': 'modified_date', '파일URL': 'file_url',
    }
    records = []
    for row in rows:
        d = map_row(headers, row, DRAW_MAP)
        if not d.get('pn') and not d.get('filename'): continue
        d['revision'] = to_num(d.get('revision'))
        d['file_size_kb'] = to_num(d.get('file_size_kb'))
        d['modified_date'] = to_date(d.get('modified_date'))
        for k in ('customer','pn','revision_status','filename','file_ext','file_url'):
            d[k] = to_text(d.get(k))
        records.append(d)
    BATCH = 500
    for i in range(0, len(records), BATCH):
        client.table('drawings').insert(records[i:i+BATCH]).execute()
    return {'table': 'drawings', 'count': len(records)}


def import_sales_ledger(client: Client, src_dir: str) -> dict:
    """매출내역_우성정밀.xlsx → sales_ledger"""
    path = os.path.join(src_dir, '매출내역_우성정밀.xlsx')
    headers, rows = read_xlsx(path, '매출내역_전표')
    records = []
    for row in rows:
        if not row[0]: continue
        d = {
            'customer': to_text(row[0]),
            'trade_subtype': to_text(row[1]),
            'voucher_no': to_int(row[2]),
            'voucher_date': to_date(row[3]),
            'item_date': to_date(row[4]),
            'item': to_text(row[5]),
            'spec': to_text(row[6]),
            'qty': to_num(row[7]),
            'unit': to_text(row[8]),
            'unit_price': to_num(row[9]),
            'amount': to_num(row[10]),
            'vat': to_num(row[11]),
            'total': to_num(row[12]),
            'remark': to_text(row[13]),
        }
        records.append(d)
    BATCH = 1000
    for i in range(0, len(records), BATCH):
        client.table('sales_ledger').insert(records[i:i+BATCH]).execute()
    return {'table': 'sales_ledger', 'count': len(records)}


def import_purchase_ledger(client: Client, src_dir: str) -> dict:
    """매입내역 3개 파일 → purchase_ledger"""
    total = 0
    for year in ('2024','2025','2026'):
        path = os.path.join(src_dir, f'{year}년_매입내역_구글시트.xlsx')
        if not os.path.exists(path): continue
        wb = openpyxl.load_workbook(path, data_only=True)
        records = []
        for sn in wb.sheetnames:
            if '합계' in sn or '요약' in sn: continue
            ws = wb[sn]
            for r in range(2, ws.max_row + 1):
                vendor = ws.cell(r, 2).value
                item = ws.cell(r, 3).value
                if not vendor or not item: continue
                records.append({
                    'trade_date': to_date(ws.cell(r, 1).value),
                    'vendor': to_text(vendor),
                    'item': to_text(item),
                    'unit': to_text(ws.cell(r, 4).value),
                    'qty': to_num(ws.cell(r, 5).value),
                    'weight': to_num(ws.cell(r, 6).value),
                    'unit_price': to_num(ws.cell(r, 7).value),
                    'amount': to_num(ws.cell(r, 8).value),
                    'vat': to_num(ws.cell(r, 9).value),
                    'total': to_num(ws.cell(r, 10).value),
                    'remark': to_text(ws.cell(r, 11).value),
                })
        BATCH = 1000
        for i in range(0, len(records), BATCH):
            client.table('purchase_ledger').insert(records[i:i+BATCH]).execute()
        total += len(records)
    return {'table': 'purchase_ledger', 'count': total}


def archive_dormant_products(client: Client) -> dict:
    """3년 이상 거래 없는 제품을 archived_at 처리"""
    threshold = (datetime.now() - timedelta(days=1095)).date().isoformat()
    # SQL로 일괄 업데이트 (raw SQL 또는 RPC)
    # supabase python client는 RPC만 지원, 또는 update with filter
    r = client.rpc('archive_dormant_products', {'threshold_date': threshold}).execute()
    return {'archived': r.data}


def run_full_import(client: Client, src_dir: str) -> list[dict]:
    """전체 import 실행 — 마스터 5종 + ledger 2종"""
    results = []
    # 1. 마스터 (참조 무결성 순서)
    results.append(import_products(client, src_dir))
    results.append(import_vendors(client, src_dir))
    results.append(import_materials(client, src_dir))
    results.append(import_drawings(client, src_dir))
    # 2. BOM (products/materials 다음)
    results.append(import_bom(client, src_dir))
    # 3. ledger
    results.append(import_sales_ledger(client, src_dir))
    results.append(import_purchase_ledger(client, src_dir))
    return results
