"""
발주서 xlsx 생성 — 완전 새 양식 (코드로 동적 생성, 템플릿 파일 의존 X)
- A4 1페이지 깔끔한 레이아웃
- 우측 발주자(우성정밀) / 좌측 수신처(거래처) 균형
- 품목 표 단정 (NO/품명/재질/규격/수량/단가/금액)
- 합계 영역 명확 (공급가액 / VAT / 합계)
- 발주 조건 (납기/지불조건/배송지/비고)
"""
import io
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins, PrintOptions


# ─── 우성정밀 자사 정보 (상수) ───
COMPANY_INFO = {
    "name": "우성정밀",
    "biz_no": "606-02-14529",
    "ceo": "김태식",
    "address": "부산광역시 기장군 산단4로 71",
    "phone": "051-527-9963",
    "fax": "051-526-6024",
    "biz_type": "제조",
    "biz_item": "자동차부품",
}

# ─── 스타일 ───
FONT_TITLE = Font(name="맑은 고딕", size=22, bold=True, color="1F3864")
FONT_HEADER = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
FONT_SUBHEADER = Font(name="맑은 고딕", size=10, bold=True, color="1F3864")
FONT_LABEL = Font(name="맑은 고딕", size=9, bold=True, color="595959")
FONT_BODY = Font(name="맑은 고딕", size=10)
FONT_SMALL = Font(name="맑은 고딕", size=9, color="595959")
FONT_TOTAL = Font(name="맑은 고딕", size=11, bold=True)

FILL_DARK = PatternFill("solid", start_color="305496")
FILL_LIGHT = PatternFill("solid", start_color="DEE6F0")
FILL_GRAY = PatternFill("solid", start_color="F2F2F2")
FILL_TOTAL = PatternFill("solid", start_color="FFF2CC")

THIN = Side(border_style="thin", color="BFBFBF")
THICK = Side(border_style="medium", color="305496")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HEADER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
RIGHT = Alignment(horizontal="right", vertical="center", indent=1)


def generate_po_number(client=None) -> str:
    """
    발주번호 생성. 동시성 안전 DB 함수 우선 호출.
    client 인자는 historical compat (현재는 사용 X). 신규 호출은
    app.services.purchase_service.generate_po_number_safe() 권장.
    """
    try:
        from app.services.purchase_service import generate_po_number_safe
        return generate_po_number_safe()
    except Exception:
        # 의존성 import 실패 시에만 Python fallback
        today = date.today()
        prefix = f"PO-{today.strftime('%Y%m')}-"
        if client is None:
            return f"{prefix}001"
        try:
            rows = client.fetch(
                "purchase_orders", "po_number",
                f"po_number=like.{prefix}*&order=po_number.desc&limit=1",
            )
            seq = int(rows[0]["po_number"].replace(prefix, "")) + 1 if rows else 1
        except Exception:
            seq = 1
        return f"{prefix}{seq:03d}"


def fill_po_template(po_data: dict, items: list[dict], vendor_info: dict = None) -> bytes:
    """
    완전 새 양식으로 발주서 xlsx 생성.
    po_data: po_number, po_date, vendor_name, delivery_date, payment_terms,
             delivery_address, contact_person, total_amount(선택), remark(선택)
    items: [{item_name, material, spec, qty, unit_price, amount}]
    vendor_info: {biz_no, ceo, address, phone} (옵션, 거래처 마스터에서 채움)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발주서"

    # ─── 페이지 설정 (A4 세로, 여백 좁게) ───
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5,
                                  header=0.3, footer=0.3)
    ws.print_options = PrintOptions(horizontalCentered=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # 컬럼 너비 (A~G 7개 컬럼: NO/품명/재질/규격/수량/단가/금액)
    # 화면 가독성과 A4 인쇄 균형 — 합 90 + scale 95%
    widths = [5, 26, 10, 14, 10, 12, 13]  # 합 90
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 인쇄 설정
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.scale = None  # fitToPage 우선
    ws.sheet_view.zoomScale = 100

    row = 1

    # ─── 1. 제목 ───
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row, 1, "발  주  서")
    cell.font = FONT_TITLE
    cell.alignment = CENTER
    ws.row_dimensions[row].height = 32
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row, 1, "PURCHASE  ORDER")
    cell.font = Font(name="Calibri", size=11, italic=True, color="595959")
    cell.alignment = CENTER
    row += 1

    # 발주번호 + 발주일
    po_no = po_data.get("po_number", "")
    po_date = po_data.get("po_date", date.today())
    if isinstance(po_date, (date, datetime)):
        po_date_str = po_date.strftime("%Y년 %m월 %d일")
    else:
        po_date_str = str(po_date)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row, 1, f"발주번호: {po_no}").font = FONT_SUBHEADER
    ws.cell(row, 1).alignment = CENTER

    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    ws.cell(row, 5, f"발주일: {po_date_str}").font = FONT_SUBHEADER
    ws.cell(row, 5).alignment = CENTER
    ws.row_dimensions[row].height = 22
    row += 2

    # ─── 2. 수신·발신 영역 ───
    # 헤더 행
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    h1 = ws.cell(row, 1, "▶ 수  신 (To)")
    h1.font = FONT_HEADER; h1.fill = FILL_DARK; h1.alignment = LEFT
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    h2 = ws.cell(row, 5, "▶ 발  신 (From)")
    h2.font = FONT_HEADER; h2.fill = FILL_DARK; h2.alignment = LEFT
    ws.row_dimensions[row].height = 20
    row += 1

    vendor_info = vendor_info or {}

    # 수신 정보 (왼쪽 A:D) / 발신 정보 (오른쪽 E:G)
    info_pairs = [
        # (label, vendor_value, woosung_value)
        ("거래처", po_data.get("vendor_name", ""), COMPANY_INFO["name"]),
        ("사업자", vendor_info.get("biz_no") or "", COMPANY_INFO["biz_no"]),
        ("대표자", vendor_info.get("ceo") or "", COMPANY_INFO["ceo"]),
        ("주  소", vendor_info.get("address") or "", COMPANY_INFO["address"]),
        ("전  화", vendor_info.get("phone") or "", COMPANY_INFO["phone"]),
        ("담당자", "", po_data.get("contact_person") or "김민수 과장 / 010-3881-1165"),
    ]

    for label, vleft, vright in info_pairs:
        # 왼쪽 라벨
        ws.cell(row, 1, label).font = FONT_LABEL
        ws.cell(row, 1).fill = FILL_GRAY
        ws.cell(row, 1).alignment = CENTER
        ws.cell(row, 1).border = BORDER_ALL
        # 왼쪽 값 (B:D 병합)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        cell = ws.cell(row, 2, vleft)
        cell.font = FONT_BODY; cell.alignment = LEFT; cell.border = BORDER_ALL
        # 오른쪽 라벨
        ws.cell(row, 5, label).font = FONT_LABEL
        ws.cell(row, 5).fill = FILL_GRAY
        ws.cell(row, 5).alignment = CENTER
        ws.cell(row, 5).border = BORDER_ALL
        # 오른쪽 값 (F:G 병합)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        cell = ws.cell(row, 6, vright)
        cell.font = FONT_BODY; cell.alignment = LEFT; cell.border = BORDER_ALL
        ws.row_dimensions[row].height = 19
        row += 1

    row += 1  # 빈 행

    # ─── 3. "아래와 같이 발주합니다" 안내 ───
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row, 1, "아래와 같이 발주합니다.")
    cell.font = Font(name="맑은 고딕", size=10, bold=True, color="305496")
    cell.alignment = LEFT
    ws.row_dimensions[row].height = 18
    row += 1

    # ─── 4. 품목 표 헤더 ───
    headers = ["NO", "품  명", "재 질", "규 격", "수 량", "단 가", "금 액"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row, ci, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_DARK
        cell.alignment = CENTER
        cell.border = BORDER_ALL
    ws.row_dimensions[row].height = 20
    row += 1

    # ─── 5. 품목 데이터 (최대 20행, 빈 행은 공백 표시) ───
    total_amount = 0
    items_to_render = items[:20] if items else []
    for i in range(20):
        item = items_to_render[i] if i < len(items_to_render) else None
        # NO
        ws.cell(row, 1, i + 1).font = FONT_BODY
        ws.cell(row, 1).alignment = CENTER
        ws.cell(row, 1).border = BORDER_ALL
        # 데이터
        if item:
            # 품명 + (메모가 있으면 줄바꿈으로 부가 표시)
            name_text = item.get("item_name", "")
            memo = item.get("memo", "")
            if memo:
                name_text = f"{name_text}\n  └ {memo}"
            ws.cell(row, 2, name_text)
            ws.cell(row, 3, item.get("material") or "")
            ws.cell(row, 4, item.get("spec") or "")
            qty = int(item.get("qty") or 0)
            up = int(item.get("unit_price") or 0)
            amt = qty * up
            ws.cell(row, 5, qty if qty else "")
            ws.cell(row, 6, up if up else "")
            ws.cell(row, 7, amt if amt else "")
            total_amount += amt
            ws.cell(row, 5).number_format = "#,##0"
            ws.cell(row, 6).number_format = '"₩"#,##0'
            ws.cell(row, 7).number_format = '"₩"#,##0'
            # 메모가 있으면 행 높이 조금 증가
            if memo:
                ws.row_dimensions[row].height = 28

        # 모든 셀 스타일
        ws.cell(row, 2).font = FONT_BODY; ws.cell(row, 2).alignment = LEFT; ws.cell(row, 2).border = BORDER_ALL
        ws.cell(row, 3).font = FONT_BODY; ws.cell(row, 3).alignment = CENTER; ws.cell(row, 3).border = BORDER_ALL
        ws.cell(row, 4).font = FONT_BODY; ws.cell(row, 4).alignment = CENTER; ws.cell(row, 4).border = BORDER_ALL
        ws.cell(row, 5).font = FONT_BODY; ws.cell(row, 5).alignment = RIGHT; ws.cell(row, 5).border = BORDER_ALL
        ws.cell(row, 6).font = FONT_BODY; ws.cell(row, 6).alignment = RIGHT; ws.cell(row, 6).border = BORDER_ALL
        ws.cell(row, 7).font = FONT_BODY; ws.cell(row, 7).alignment = RIGHT; ws.cell(row, 7).border = BORDER_ALL

        # 짝수 행 옅은 회색
        if i % 2 == 1:
            for ci in range(1, 8):
                ws.cell(row, ci).fill = PatternFill("solid", start_color="FAFAFA")

        ws.row_dimensions[row].height = 19
        row += 1

    # ─── 6. 합계 영역 ───
    vat = int(total_amount * 0.1)
    grand_total = total_amount + vat

    summary_rows = [
        ("공급가액 합계", total_amount, FILL_LIGHT),
        ("부 가 세 (10%)", vat, FILL_LIGHT),
        ("총 합 계", grand_total, FILL_TOTAL),
    ]
    for label, value, fill in summary_rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row, 1, label)
        cell.font = FONT_TOTAL; cell.alignment = RIGHT; cell.fill = fill
        cell.border = BORDER_ALL
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        cell = ws.cell(row, 6, value)
        cell.font = FONT_TOTAL; cell.alignment = RIGHT; cell.fill = fill
        cell.border = BORDER_ALL
        cell.number_format = '"₩"#,##0'
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1

    # ─── 7. 발주 조건 ───
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row, 1, "◆ 발주 조건")
    cell.font = FONT_SUBHEADER
    cell.fill = FILL_LIGHT
    cell.alignment = LEFT
    ws.row_dimensions[row].height = 22
    row += 1

    conditions = [
        ("• 납  기", po_data.get("delivery_date") or ""),
        ("• 지불조건", po_data.get("payment_terms") or "말일 마감 60일 현금"),
        ("• 배 송 지", po_data.get("delivery_address") or COMPANY_INFO["address"]),
        ("• 비  고", po_data.get("remark") or ""),
    ]
    for label, value in conditions:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row, 1, label)
        cell.font = FONT_LABEL; cell.alignment = LEFT; cell.fill = FILL_GRAY
        cell.border = BORDER_ALL
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        cell = ws.cell(row, 3, value)
        cell.font = FONT_BODY; cell.alignment = LEFT
        cell.border = BORDER_ALL
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1

    # ─── 8. 푸터 (회사 인장 영역) ───
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row, 1,
                   f"{COMPANY_INFO['name']}  ·  {COMPANY_INFO['address']}  ·  "
                   f"Tel {COMPANY_INFO['phone']}  ·  Fax {COMPANY_INFO['fax']}")
    cell.font = FONT_SMALL
    cell.alignment = CENTER
    ws.row_dimensions[row].height = 18
    row += 1

    # ─── 저장 ───
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
