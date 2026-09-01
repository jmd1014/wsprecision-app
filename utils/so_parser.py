"""
수주 파일 자동 파싱 (3개 거래처 양식)
- HDX 엑셀 (53열, 자재코드는 우성정밀 품번 형식 — 거의 자동 매칭)
- 미진정밀 외주발주품목조회 엑셀 (37열, 제품번호=우성정밀 품번 100%)
- 엠제이티 PDF (단순 표 1개)
"""
import io
import re
from datetime import date, datetime
import openpyxl


def _to_num(v):
    if v is None or v == "": return None
    try: return float(str(v).replace(",", ""))
    except: return None


def _to_int(v):
    n = _to_num(v)
    return int(n) if n is not None else None


def _to_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    s = str(v).strip()
    if not s: return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%Y%m%d"):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    return None


def parse_hdx_excel(file_bytes: bytes, filename: str = "") -> list[dict]:
    """
    HDX 수주 엑셀 파싱 — 행이 곧 품목.
    같은 수주번호가 여러 행 (= 한 수주에 여러 품목)일 수 있어 그룹핑은 호출자가 처리.

    반환: 각 품목별 dict 리스트 (so_number, line_no, item, qty, price 등)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]

    def col(name):
        try: return headers.index(name) + 1
        except: return None

    items = []
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, col("No")).value if col("No") else None
        if no is None: continue
        item = {
            "_source": "HDX_EXCEL",
            "_raw_filename": filename,
            "customer": "HDX",
            "so_number": str(ws.cell(r, col("수주번호")).value) if col("수주번호") else "",
            "line_no": int(ws.cell(r, col("항번")).value) if col("항번") and ws.cell(r, col("항번")).value else int(no),
            "so_date": _to_date(ws.cell(r, col("수주일자")).value) if col("수주일자") else None,
            "due_date": _to_date(ws.cell(r, col("납기요청일")).value) if col("납기요청일") else None,
            "customer_part_no": ws.cell(r, col("자재코드")).value if col("자재코드") else None,
            "customer_item_name": ws.cell(r, col("자재명")).value if col("자재명") else None,
            "qty": _to_num(ws.cell(r, col("수량")).value) if col("수량") else None,
            "unit": ws.cell(r, col("단위")).value if col("단위") else "EA",
            "unit_price": _to_num(ws.cell(r, col("단가")).value) if col("단가") else None,
            "amount": _to_num(ws.cell(r, col("금액")).value) if col("금액") else None,
            "delivery_address": ws.cell(r, col("납품처 장소")).value if col("납품처 장소") else None,
            "remark": ws.cell(r, col("특이사항")).value if col("특이사항") else None,
            "status_raw": ws.cell(r, col("입고상태")).value if col("입고상태") else None,
            "raw_row": {h: ws.cell(r, ci+1).value for ci, h in enumerate(headers) if h},
        }
        items.append(item)
    return items


def parse_mijin_excel(file_bytes: bytes, filename: str = "") -> list[dict]:
    """
    미진정밀 외주발주품목조회 엑셀 파싱.
    행 1 = 타이틀 ("외주발주품목조회"), 행 2 = 헤더, 행 3 = TOTAL, 행 4부터 데이터.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[2]]  # 행 2가 헤더

    def col(name):
        try: return headers.index(name) + 1
        except: return None

    items = []
    line_seq = {}  # so_number → 다음 line_no
    for r in range(4, ws.max_row + 1):  # 행 3은 TOTAL, 행 4부터
        sel = ws.cell(r, 1).value  # 선택 컬럼
        # TOTAL 행 또는 빈 행 skip
        if sel is None: continue
        po = ws.cell(r, col("외주발주번호")).value
        if not po or str(po).strip() in ("TOTAL", ""): continue

        so_num = str(po)
        line_seq[so_num] = line_seq.get(so_num, 0) + 1

        item = {
            "_source": "MIJIN_EXCEL",
            "_raw_filename": filename,
            "customer": "미진정밀",
            "so_number": so_num,
            "line_no": line_seq[so_num],
            "so_date": _to_date(ws.cell(r, col("외주발주일")).value) if col("외주발주일") else None,
            "due_date": _to_date(ws.cell(r, col("납기일")).value) if col("납기일") else None,
            # 미진은 제품번호가 우성정밀 품번
            "customer_part_no": ws.cell(r, col("제품번호")).value if col("제품번호") else None,
            "canonical_pn_hint": ws.cell(r, col("제품번호")).value if col("제품번호") else None,
            "customer_item_name": ws.cell(r, col("제품명")).value if col("제품명") else None,
            "qty": _to_num(ws.cell(r, col("발주수량")).value) if col("발주수량") else None,
            "received_qty": _to_num(ws.cell(r, col("납품수량")).value) if col("납품수량") else 0,
            "unit_price": _to_num(ws.cell(r, col("단가")).value) if col("단가") else None,
            "amount": _to_num(ws.cell(r, col("금액")).value) if col("금액") else None,
            "vat": _to_num(ws.cell(r, col("부가세")).value) if col("부가세") else None,
            "total": _to_num(ws.cell(r, col("금액계")).value) if col("금액계") else None,
            "mes_work_order": ws.cell(r, col("작업지시번호")).value if col("작업지시번호") else None,
            "status_raw": ws.cell(r, col("진행상태")).value if col("진행상태") else None,
            "raw_row": {h: ws.cell(r, ci+1).value for ci, h in enumerate(headers) if h},
        }
        items.append(item)
    return items


def _pick_line_nums(raw_nums: list[str]) -> list:
    """품목 행의 숫자 토큰 해석 — '수량 단가 금액' 검산 기반.

    PDF 가 "78,000,000"을 "7 8,000,000"으로 쪼개는 경우가 있어 병합
    후보를 만들되, 무조건 병합하면 정상 라인("100 6,000 600,000")까지
    이어붙는 사고가 난다(2026-08-25 MJT-PO26-우성-708). 게다가 한
    행에서 수량·금액이 동시에 쪼개지기도 한다("3 ,600 6,000 2
    1,600,000" = 3,600 × 6,000 = 21,600,000 — 2026-09-01 우성-725)
    — 겹치지 않는 병합 위치의 모든 조합을 후보로 만들고 수량 × 단가
    = 금액이 성립하는 해석을 고른다. 없으면 그대로 읽는다."""
    from itertools import combinations

    positions = [i for i in range(len(raw_nums) - 1)
                 if ',' not in raw_nums[i] and ',' in raw_nums[i + 1]]
    subsets = [()]
    for _r in range(1, len(positions) + 1):
        for combo in combinations(positions, _r):
            # 인접 위치 동시 병합은 토큰이 겹쳐 무효
            if all(b - a >= 2 for a, b in zip(combo, combo[1:])):
                subsets.append(combo)
    cands = []
    for combo in subsets:
        toks = list(raw_nums)
        for i in sorted(combo, reverse=True):
            toks[i:i + 2] = [toks[i] + toks[i + 1].replace(',', '')]
        cands.append([_to_int(t) for t in toks])
    good = [c for c in cands if len(c) >= 3 and c[0] and c[1]
            and c[0] * c[1] == c[2]]
    if good:
        # 숫자 3개(수량·단가·금액)로 딱 떨어지는 해석 우선
        return sorted(good, key=len)[0]
    return cands[0]


def parse_mjt_pdf(file_bytes: bytes, filename: str = "") -> list[dict]:
    """
    엠제이티(MJT) PDF 발주서 파싱 — 텍스트 라인 기반 정규식
    표 추출 대신 텍스트 직접 매칭 (PDF 표 셀 분리가 불안정함)
    """
    try:
        import pdfplumber
    except ImportError:
        raise RuntimeError("pdfplumber 필요")

    items = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        full_text = "\n".join(pg.extract_text() or "" for pg in pdf.pages)

        # 헤더 정보 추출
        m_no = re.search(r'발주번호\s+([A-Z0-9가-힣\-]+)', full_text)
        so_number = m_no.group(1) if m_no else ""
        m_date = re.search(r'발주일자\s+(\d{4}-\d{2}-\d{2})', full_text)
        so_date = _to_date(m_date.group(1)) if m_date else None
        m_due = re.search(r'납기일자\s+(\d{4}-\d{2}-\d{2})', full_text)
        due_date = _to_date(m_due.group(1)) if m_due else None

        # 라인 파싱 — 품목 행 패턴
        # 예: "1 12HFDVN-VM-03 EA 13,000 6,000 78,000,000 06/05"
        # 또는 PDF 렌더링으로 "1 12HFDVN-VM-03 EA 13,000 6,000 7 8,000,000 06/05"
        # 핵심: 첫 토큰 = 라인번호(숫자), 두번째 = 품번(공백 X), 세번째 = 단위, 나머지 = 숫자들
        for line in full_text.split("\n"):
            line = line.strip()
            if not line: continue
            # 라인 시작이 숫자 (1~99) + 공백 + 품번 + 단위
            m = re.match(
                r'^(\d{1,3})\s+'                    # NO
                r'([A-Za-z0-9][A-Za-z0-9\-_/;\.]+)\s+' # 품번
                r'(EA|개|kg|KG|m|M|set|SET|L|ℓ)\s+'  # 단위
                r'(.+)$',                            # 나머지 (수량, 단가, 금액, 비고)
                line
            )
            if not m: continue
            line_no = int(m.group(1))
            pn = m.group(2).strip()
            unit = m.group(3).strip().upper()
            rest = m.group(4).strip()

            # rest에서 숫자 추출 — "수량 단가 금액" 이 정석이지만 PDF 가
            # "78,000,000"을 "7 8,000,000"으로 쪼개는 경우가 있다.
            # 무조건 병합하면 정상 라인("100 6,000 600,000")까지
            # 이어붙는 사고(2026-08-25 MJT-PO26-우성-708)가 나므로,
            # 해석 후보를 만들고 '수량 × 단가 = 금액' 검산으로 고른다.
            tokens = rest.split()
            raw_nums = [t for t in tokens if re.match(r'^[\d,]+$', t)]
            remark_parts = [t for t in tokens
                            if not re.match(r'^[\d,]+$', t)]

            num_only = _pick_line_nums(raw_nums)

            qty = num_only[0] if num_only else None
            unit_price = num_only[1] if len(num_only) >= 2 else None
            amount = num_only[2] if len(num_only) >= 3 else (
                qty * unit_price if qty and unit_price else None
            )
            remark = " ".join(remark_parts) if remark_parts else ""

            # 헤더 납기일자가 비어 있는 양식 — 라인 비고의 MM/DD 를
            # 납기로 해석 (발주일보다 앞이면 이듬해)
            _line_due = due_date
            if _line_due is None and so_date:
                _md9 = re.search(r'\b(\d{2})/(\d{2})\b', remark)
                if _md9:
                    try:
                        _yy = so_date.year
                        _cand = date(_yy, int(_md9.group(1)),
                                     int(_md9.group(2)))
                        if _cand < so_date:
                            _cand = date(_yy + 1, int(_md9.group(1)),
                                         int(_md9.group(2)))
                        _line_due = _cand
                    except ValueError:
                        pass

            items.append({
                "_source": "MJT_PDF",
                "_raw_filename": filename,
                "customer": "(주)엠제이티",
                "so_number": so_number,
                "line_no": line_no,
                "so_date": so_date,
                "due_date": _line_due,
                "customer_part_no": pn,
                "canonical_pn_hint": pn,
                "customer_item_name": "",
                "qty": qty,
                "unit": unit,
                "unit_price": unit_price,
                "amount": amount,
                "remark": remark,
                "raw_row": {"line": line_no, "pn": pn, "raw_line": line},
            })

    return items


# ────────────────────────────────────────────────
# 양식 자동 인식
# ────────────────────────────────────────────────

def detect_so_format(file_bytes: bytes, filename: str = "") -> str:
    """
    파일 양식 자동 인식. 반환: 'HDX' / 'MIJIN' / 'MJT_PDF' / 'UNKNOWN'
    """
    fname_lower = filename.lower()
    # PDF
    if fname_lower.endswith('.pdf'):
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                text = "\n".join((pg.extract_text() or "")[:2000] for pg in pdf.pages[:1])
            if any(k in text for k in ['MJT', '엠제이티', 'mjt-global', 'MJT-PO']):
                return 'MJT_PDF'
            # 향후 HDX/DIC PDF 양식 추가 가능
            return 'UNKNOWN_PDF'
        except Exception:
            return 'UNKNOWN_PDF'

    # 엑셀
    if not (fname_lower.endswith('.xlsx') or fname_lower.endswith('.xls')):
        return 'UNKNOWN'
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        # 행 1, 2, 3 헤더/타이틀 검사
        cells_top = []
        for r in range(1, min(4, ws.max_row + 1)):
            for c in range(1, min(15, ws.max_column + 1)):
                v = ws.cell(r, c).value
                if v is not None: cells_top.append(str(v))
        text_block = " | ".join(cells_top)

        # 미진정밀: 첫 행 "외주발주품목조회" 타이틀
        if '외주발주품목조회' in text_block:
            return 'MIJIN'
        # HDX: "MRP", "수주번호", "업체자재코드", "납기요청일" 등 동시 등장
        hdx_keys = ['수주번호', '업체자재코드', '협력업체', 'MRP']
        if sum(1 for k in hdx_keys if k in text_block) >= 2:
            return 'HDX'
        # 거래처명을 파일명으로 추정
        if 'HDX' in fname_lower.upper() or 'hdx' in fname_lower:
            return 'HDX'
        if '미진' in filename or 'mijin' in fname_lower:
            return 'MIJIN'
        return 'UNKNOWN_EXCEL'
    except Exception:
        return 'UNKNOWN'


def parse_so_auto(file_bytes: bytes, filename: str = "") -> tuple[str, list[dict]]:
    """자동 양식 인식 후 적절한 파서 호출. (format, items) 반환"""
    fmt = detect_so_format(file_bytes, filename)
    if fmt == 'HDX':
        return fmt, parse_hdx_excel(file_bytes, filename)
    if fmt == 'MIJIN':
        return fmt, parse_mijin_excel(file_bytes, filename)
    if fmt == 'MJT_PDF':
        return fmt, parse_mjt_pdf(file_bytes, filename)
    return fmt, []


# ─── 우성정밀 품번 매칭 ───
def match_canonical_pn(items: list[dict], canonical_map: dict) -> list[dict]:
    """
    파싱된 items에 우성정밀 품번/product_id 자동 매칭.
    canonical_map: { match_key: canonical_pn } (db.fetch로 product_master에서 빌드)
    """
    def mk(s):
        if not s: return ""
        s = str(s).upper()
        s = re.sub(r'\([^)]*\)', '', s)
        s = re.sub(r'[\s\-_·,\.]+', '', s)
        return s

    def normalize_item(item_text):
        """;OP, ;PM 등 제거"""
        if not item_text: return ""
        pn = str(item_text).strip().split(';')[0].strip()
        return pn

    for it in items:
        candidates = [
            it.get("canonical_pn_hint"),
            it.get("customer_part_no"),
        ]
        matched = None
        for c in candidates:
            if not c: continue
            # 1) 정확 매칭
            if mk(c) in canonical_map:
                matched = canonical_map[mk(c)]; break
            # 2) ;XXX 제거 후 매칭
            norm = normalize_item(c)
            if mk(norm) in canonical_map:
                matched = canonical_map[mk(norm)]; break
        it["matched_pn"] = matched
    return items


def parse_so_file(customer_type: str, file_bytes: bytes, filename: str = "") -> list[dict]:
    """통합 진입점"""
    if customer_type == "HDX":
        return parse_hdx_excel(file_bytes, filename)
    elif customer_type == "미진정밀":
        return parse_mijin_excel(file_bytes, filename)
    elif customer_type in ("㈜엠제이티", "엠제이티"):
        return parse_mjt_pdf(file_bytes, filename)
    else:
        raise ValueError(f"지원하지 않는 양식: {customer_type}")


def group_by_so_number(items: list[dict]) -> list[dict]:
    """
    같은 so_number 행들을 그룹핑하여 sales_orders 헤더 + items로 변환.
    반환: [{header: {...}, items: [{...}, ...]}, ...]
    """
    groups = {}
    for it in items:
        so = it.get("so_number") or "(no_number)"
        if so not in groups:
            groups[so] = {"header": None, "items": []}
        groups[so]["items"].append(it)

    result = []
    for so, g in groups.items():
        items_list = g["items"]
        first = items_list[0]
        header = {
            "so_number": so,
            "customer": first["customer"],
            "so_date": first.get("so_date"),
            "due_date": min((i["due_date"] for i in items_list if i.get("due_date")), default=None),
            "total_amount": sum((i.get("amount") or 0) for i in items_list),
            "vat": sum((i.get("vat") or (i.get("amount") or 0) * 0.1) for i in items_list),
            "source": first.get("_source"),
            "source_file": first.get("_raw_filename"),
            "delivery_address": first.get("delivery_address"),
            "status": "DRAFT",
        }
        result.append({"header": header, "items": items_list})
    return result
